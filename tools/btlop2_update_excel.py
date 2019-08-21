##
#	@file	btlop2_update_excel.py
#	@brief	Update weapon and skill sheet by MS sheet
#

import	os
import	sys
import	openpyxl

# ----------------------------------------------------------------------
#	Constants
SHEETNAME_MS        = 'MS'
SHEETNAME_WEAPON1   = 'Weapon1'
SHEETNAME_WEAPON2   = 'Weapon2'
SHEETNAME_SUBWEAPON = 'SubWeapon'
SHEETNAME_SKILL     = 'Skill'

# ----------------------------------------------------------------------
##	@brief	Search column
def search_col(sheet, name):
	i = sheet.min_column
	while i <= sheet.max_column:
		val = sheet.cell( row=sheet.min_row, column=i ).value
		if val:
			val = str(val).strip()
			if val == name:
				return i
		i += 1
	return -1

# ----------------------------------------------------------------------
##	@brief	Collect items
def collect_items(sheet, colname):
	temp = {}
	name_col = search_col( sheet, 'name' )
	col = search_col( sheet, colname )
	if col >= 0:
		i = sheet.min_row + 1
		while i <= sheet.max_row:
			val = sheet.cell( row=i, column=col ).value
			if type(val) is str:
				items = val.split('\n')
				for item in items:
					itemname = item.strip()
					if itemname == '':
						continue
					if itemname not in temp:
						temp[itemname] = []

					name = str(sheet.cell( row=i, column=name_col ).value).strip()
					if name != '':
						temp[itemname].append( name )
			i += 1
	return temp

# ----------------------------------------------------------------------
##	@brief	Collect items for sub-weapon
def collect_sub_items(sheet, colname):
	ret = {}
	ret2 = {}
	name_col = search_col( sheet, 'name' )
	level_col = search_col( sheet, 'level' )
	col = search_col( sheet, colname )
	if col >= 0:
		i = sheet.min_row + 1
		while i <= sheet.max_row:
			name = sheet.cell( row=i, column=name_col ).value
			if type(name) is str:
				if name != '':
					level = sheet.cell( row=i, column=level_col ).value
					if type(level) is str or type(level) is float:
						level = int(level)
					elif type(level) is not int:
						level = 1
					val = sheet.cell( row=i, column=col ).value
					if type(val) is str:
						items = val.split('\n')
						temp = []
						for item in items:
							itemname = item.strip()
							if itemname == '':
								continue
							temp.append( itemname )
						if len(temp) > 0:
							ret[name] = sorted( temp )
							if name in ret2:
								if ret2[name] < level:
									ret2[name] = level
							else:
								ret2[name] = level
			i += 1
	return ret, ret2

# ----------------------------------------------------------------------
##	@brief	Update items
def update_items(sheet, items):
	name_col = sheet.min_column
	i0 = sheet.min_row + 1
	i1 = 0
	while i1 < len(items):
		val = sheet.cell( row=i0, column=name_col ).value
		if val:
			val = str(val).strip()
		else:
			val = ''
		if val == '' or val > items[i1]:
			sheet.insert_rows( i0 )
			sheet.cell( row=i0, column=name_col, value=items[i1] )
			i0 += 1
			i1 += 1
		elif val < items[i1]:
			i0 += 1
		else:
			i0 += 1
			i1 += 1

# ----------------------------------------------------------------------
##	@brief	Update items for sub-weapon
def update_sub_items(sheet, items, levels):
	body_list = sorted( items )
	item_list = []
	for body in body_list:
		for name in items[body]:
			level = levels[body]
			while level >= 1:
				item_list.append( [body, name, level] )
				level -= 1

	body_col = search_col( sheet, 'body' )
	name_col = search_col( sheet, 'name' )
	level_col = search_col( sheet, 'level' )
	i0 = sheet.min_row + 1
	i1 = 0
	less = None		# True if [i0] less than [i1]
	while i1 < len(item_list):
		body = sheet.cell( row=i0, column=body_col ).value
		if type(body) is str:
			body = body.strip()
		else:
			body = ''
		if body == '':
			less = True
		elif body < item_list[i1][0]:
			less = True
		elif body > item_list[i1][0]:
			less = False
		else:
			val = sheet.cell( row=i0, column=name_col ).value
			if type(val) is str:
				val = val.strip()
			else:
				val = ''
			if val < item_list[i1][1]:
				less = True
			elif val > item_list[i1][1]:
				less = False
			else:
				level = sheet.cell( row=i0, column=level_col ).value
				if type(level) is str or type(level) is float:
					level = init(level)
				elif type(level) is not int:
					level = 1
				if level > item_list[i1][2]:
					less = True
				elif level < item_list[i1][2]:
					less = False
				else:
					less = None
		if less is None:
			pass
		elif less:
			sheet.insert_rows( i0 )
			sheet.cell( row=i0, column=body_col, value=item_list[i1][0] )
			sheet.cell( row=i0, column=name_col, value=item_list[i1][1] )
			sheet.cell( row=i0, column=level_col, value=item_list[i1][2] )
			i0 += 1
			i1 += 1
		elif not less:
			i0 += 1
		else:
			i0 += 1
			i1 += 1

# ----------------------------------------------------------------------
#	Main

argv = sys.argv
argc = len(argv)

if argc < 2:
	sys.stdout.write( 'Error: No input file' )
	sys.exit( 1 )

filename = argv[1]
if not os.path.isfile(filename):
	sys.stdout.write( f'Error: File not found : {filename}' )
	sys.exit( 1 )

book = openpyxl.load_workbook( filename )

sheet_ms = book[SHEETNAME_MS]
sheet_weapon1 = book[SHEETNAME_WEAPON1]

# Update main weapon (range) list
weapon1 = collect_items( sheet_ms, 'main_weapon1' )
update_items( book[SHEETNAME_WEAPON1], sorted(weapon1) )

# Update main weapon (melee) list
weapon2 = collect_items( sheet_ms, 'main_weapon2' )
update_items( book[SHEETNAME_WEAPON2], sorted(weapon2) )

# Update subweapon list
subweapon, level = collect_sub_items( sheet_ms, 'sub_weapon' )
subweapon2, level2 = collect_sub_items( sheet_weapon1, 'sub_weapon' )
subweapon.update( subweapon2 )
level.update( level2 )
update_sub_items( book[SHEETNAME_SUBWEAPON], subweapon, level )

# Update skill list
skills = collect_items( sheet_ms, 'skills' )
update_items( book[SHEETNAME_SKILL], sorted(skills) )

# Save to file
book.save( filename )
sys.exit( 0 )
